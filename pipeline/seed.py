"""
pipeline/seed.py — Seed step: WFS fetch + Eckdaten XLSX merge.

Produces: data/schools_index.yaml
Run via: just seed  (or: uv run python pipeline/seed.py)
"""
from __future__ import annotations

import asyncio
import io
import logging
from pathlib import Path
from typing import Any

import httpx
import openpyxl
from ruamel.yaml import YAML

logger = logging.getLogger(__name__)

WFS_URL = "https://gdi.berlin.de/services/wfs/schulen"
ECKDATEN_URL = "https://www.berlin.de/sen/bildung/service/daten/od-eckdaten-allg-2024.xlsx"


async def fetch_gymnasien(include_private: bool = True) -> list[dict]:
    """Fetch all Berlin Gymnasien from GDI Berlin WFS. Returns list of school dicts."""
    params = {
        "service": "WFS",
        "version": "2.0.0",
        "request": "GetFeature",
        "typeNames": "schulen",
        "outputFormat": "application/json",
        "srsName": "EPSG:4326",           # CRITICAL: always include — prevents EPSG:25833 default
        "CQL_FILTER": "schulart='Gymnasium'",
    }
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(WFS_URL, params=params)
        resp.raise_for_status()
        data = resp.json()

    schools = []
    for feature in data["features"]:
        props = feature["properties"]
        traeger = props.get("traeger", "")
        if not include_private and traeger != "öffentlich":
            continue
        # GeoJSON coordinates are [lng, lat]
        lng, lat = feature["geometry"]["coordinates"]
        hausnr = (props.get("hausnr") or "").strip()
        address = f"{props.get('strasse', '')} {hausnr}, {props.get('plz', '')} Berlin".strip()
        schools.append({
            "school_id": str(props["bsn"]),
            "name": (props.get("schulname") or "").strip(),
            "district": props.get("bezirk"),
            "address": address or None,
            "coords": {"lat": lat, "lng": lng},
            "website": props.get("internet") or None,
            "phone": props.get("telefon") or None,
            "email": props.get("email") or None,
            "traeger": traeger or None,
            "student_count": None,   # filled by XLSX merge below
            "teacher_count": None,
        })
    return schools


async def fetch_eckdaten() -> dict[str, dict]:
    """
    Fetch Eckdaten XLSX and return dict keyed by BSN (str) with student/teacher counts.
    Returns empty dict if fetch fails (non-blocking — XLSX is optional data).
    """
    async with httpx.AsyncClient(timeout=30) as client:
        try:
            resp = await client.get(ECKDATEN_URL)
            resp.raise_for_status()
        except httpx.HTTPError as e:
            logger.warning("Eckdaten XLSX fetch failed: %s — student/teacher counts will be null", e)
            return {}

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Find BSN, student count, and teacher count column indices
    # Column names vary — check multiple possibilities
    bsn_col = next((i for i, h in enumerate(headers) if h in ("BSN", "Schulnummer", "bsn")), None)
    student_col = next((i for i, h in enumerate(headers) if "Schüler" in h or "schueler" in h.lower() or "SuS" in h), None)
    teacher_col = next((i for i, h in enumerate(headers) if "Lehrkräfte" in h or "Lehrer" in h or "lehrkraefte" in h.lower()), None)

    if bsn_col is None:
        logger.warning("BSN column not found in Eckdaten XLSX. Headers: %s", headers[:10])
        return {}

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        bsn = row[bsn_col]
        if bsn is None:
            continue
        entry: dict[str, Any] = {}
        if student_col is not None and row[student_col] is not None:
            try:
                entry["student_count"] = int(row[student_col])
            except (ValueError, TypeError):
                pass
        if teacher_col is not None and row[teacher_col] is not None:
            try:
                entry["teacher_count"] = int(row[teacher_col])
            except (ValueError, TypeError):
                pass
        result[str(bsn).strip()] = entry
    return result


def merge_eckdaten(schools: list[dict], eckdaten: dict[str, dict]) -> list[dict]:
    """Merge student/teacher counts into school list by BSN (school_id)."""
    merged = 0
    for school in schools:
        sid = school["school_id"]
        if sid in eckdaten:
            school["student_count"] = eckdaten[sid].get("student_count")
            school["teacher_count"] = eckdaten[sid].get("teacher_count")
            merged += 1
    logger.info("Eckdaten merge: %d/%d schools matched by BSN", merged, len(schools))
    return schools


def write_index(schools: list[dict], index_path: Path) -> None:
    """Write schools list to YAML index file."""
    index_path.parent.mkdir(parents=True, exist_ok=True)
    yaml = YAML()
    yaml.default_flow_style = False
    yaml.width = 200
    with index_path.open("w", encoding="utf-8") as f:
        yaml.dump(schools, f)


async def run_seed(config: dict) -> list[dict]:
    """
    Run the seed step. Returns list of school dicts.
    Called by pipeline/run.py orchestrator.
    """
    include_private = config.get("include_private_schools", True)
    index_path = Path(config.get("index_file", "data/schools_index.yaml"))

    logger.info("Fetching Gymnasien from WFS (include_private=%s)...", include_private)
    schools = await fetch_gymnasien(include_private=include_private)
    logger.info("WFS returned %d schools", len(schools))

    # Validate coordinates are WGS84 (Pitfall 1 detection)
    bad_coords = [s for s in schools if s["coords"]["lat"] > 1000 or s["coords"]["lng"] > 1000]
    if bad_coords:
        raise ValueError(
            f"WFS returned non-WGS84 coordinates for {len(bad_coords)} schools. "
            "Check srsName=EPSG:4326 parameter."
        )

    logger.info("Fetching Eckdaten XLSX for student/teacher counts...")
    eckdaten = await fetch_eckdaten()
    schools = merge_eckdaten(schools, eckdaten)

    write_index(schools, index_path)
    logger.info("Written schools_index.yaml with %d schools → %s", len(schools), index_path)

    private_count = sum(1 for s in schools if s.get("traeger") != "öffentlich")
    public_count = len(schools) - private_count
    print(f"Seed complete: {len(schools)} schools ({public_count} public, {private_count} private)")
    print(f"Eckdaten merge: {sum(1 for s in schools if s['student_count'])} schools with counts")
    print(f"Output: {index_path}")

    return schools


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    # Load config from pipeline.yaml
    yaml = YAML()
    config_path = Path("pipeline/pipeline.yaml")
    if not config_path.exists():
        config_path = Path("pipeline.yaml")
    config = yaml.load(config_path) if config_path.exists() else {}
    asyncio.run(run_seed(config))
    sys.exit(0)
